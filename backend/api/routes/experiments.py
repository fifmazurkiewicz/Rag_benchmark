from __future__ import annotations

import hashlib
import json
import logging
import pathlib
import uuid
from fastapi import APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse

from backend.models.experiment import ExperimentConfig, RunStatus
from backend.models.result import ExperimentResult

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/experiments", tags=["experiments"])

EXPERIMENTS_DIR = pathlib.Path("experiments")
EXPERIMENTS_DIR.mkdir(exist_ok=True)
RESULTS_DIR = pathlib.Path("results")
RESULTS_DIR.mkdir(exist_ok=True)
CACHE_FILE = RESULTS_DIR / "cache.json"

_active_runs: dict[str, RunStatus] = {}


# ---------------------------------------------------------------------------
# Cache helpers
# ---------------------------------------------------------------------------

def _config_hash(config: ExperimentConfig) -> str:
    """Stable SHA-256 hash of dataset + sorted pipeline configs (name ignored)."""
    pipelines = [
        {k: v for k, v in sorted(p.model_dump().items()) if k != "name"}
        for p in config.pipelines
    ]
    key = json.dumps(
        {"dataset": config.dataset, "pipelines": pipelines, "metrics": sorted(config.metrics)},
        sort_keys=True,
    )
    return hashlib.sha256(key.encode()).hexdigest()[:16]


def _load_cache() -> dict[str, str]:
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text())
        except Exception:
            pass
    return {}


def _save_cache(cache: dict[str, str]) -> None:
    CACHE_FILE.write_text(json.dumps(cache, indent=2))


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/", response_model=list[str])
def list_experiments():
    return [p.stem for p in EXPERIMENTS_DIR.glob("*.json")]


@router.get("/{name}")
def get_experiment(name: str):
    path = EXPERIMENTS_DIR / f"{name}.json"
    if not path.exists():
        raise HTTPException(404, f"Experiment '{name}' not found")
    return json.loads(path.read_text())


@router.post("/")
def save_experiment(config: ExperimentConfig):
    path = EXPERIMENTS_DIR / f"{config.name}.json"
    path.write_text(config.model_dump_json(indent=2))
    return {"saved": config.name}


@router.post("/{name}/run")
def run_experiment(name: str, background_tasks: BackgroundTasks, force: bool = False):
    path = EXPERIMENTS_DIR / f"{name}.json"
    if not path.exists():
        raise HTTPException(404, f"Experiment '{name}' not found")
    config = ExperimentConfig.model_validate_json(path.read_text())

    if not force:
        cache = _load_cache()
        h = _config_hash(config)
        if h in cache:
            cached_run_id = cache[h]
            if (RESULTS_DIR / f"{cached_run_id}.json").exists():
                return {"run_id": cached_run_id, "cached": True}

    run_id = str(uuid.uuid4())[:8]
    status = RunStatus(run_id=run_id, experiment_name=name, status="pending")
    _active_runs[run_id] = status
    background_tasks.add_task(_execute_run, run_id, config)
    return {"run_id": run_id, "cached": False}


@router.get("/runs/{run_id}", response_model=RunStatus)
def get_run_status(run_id: str):
    if run_id not in _active_runs:
        raise HTTPException(404, f"Run '{run_id}' not found")
    return _active_runs[run_id]


@router.get("/results/")
def list_results():
    return [p.stem for p in RESULTS_DIR.glob("*.json") if p.stem != "cache"]


@router.get("/results/{run_id}")
def get_run_result(run_id: str):
    path = RESULTS_DIR / f"{run_id}.json"
    if not path.exists():
        raise HTTPException(404, f"Results for run '{run_id}' not found")
    return json.loads(path.read_text())


@router.get("/results/{run_id}/export")
def export_run_excel(run_id: str):
    """Download run results as Excel (.xlsx) — Summary sheet + one sheet per pipeline."""
    path = RESULTS_DIR / f"{run_id}.json"
    if not path.exists():
        raise HTTPException(404, f"Results for run '{run_id}' not found")
    data = json.loads(path.read_text())
    xlsx_bytes = _build_excel(data)
    filename = f"rag_results_{run_id}.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _build_excel(data: dict) -> bytes:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    DARK = PatternFill("solid", fgColor="1E293B")
    BOLD_LIGHT = Font(bold=True, color="E2E8F0")

    def header(ws, headers: list[str], row: int = 1) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = BOLD_LIGHT
            cell.fill = DARK
            cell.alignment = Alignment(horizontal="center")

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    pipeline_results = data.get("pipeline_results", [])
    all_metrics = list({m["name"] for pr in pipeline_results for m in pr.get("metrics", [])})

    header(ws_sum, ["Pipeline", "Type", "Latency avg (ms)", "Tokens"] + all_metrics)
    for row, pr in enumerate(pipeline_results, 2):
        ws_sum.cell(row=row, column=1, value=pr.get("pipeline_name", ""))
        ws_sum.cell(row=row, column=2, value=pr.get("pipeline_type", ""))
        ws_sum.cell(row=row, column=3, value=round(pr.get("avg_latency_ms", 0), 1))
        ws_sum.cell(row=row, column=4, value=pr.get("total_tokens", 0))
        metric_map = {m["name"]: m["value"] for m in pr.get("metrics", [])}
        for col, mn in enumerate(all_metrics, 5):
            val = metric_map.get(mn)
            cell = ws_sum.cell(row=row, column=col, value=round(val, 4) if val is not None and val >= 0 else None)
            if val is not None and val >= 0:
                color = "14532D" if val >= 0.7 else "713F12" if val >= 0.4 else "7F1D1D"
                cell.fill = PatternFill("solid", fgColor=color)

    for col in range(1, len(all_metrics) + 5):
        ws_sum.column_dimensions[get_column_letter(col)].width = 20

    # ── One sheet per pipeline (answers) ──────────────────────────────────
    for pr in pipeline_results:
        ws = wb.create_sheet(title=pr.get("pipeline_name", "pipeline")[:31])
        header(ws, ["#", "Question", "Ground truth", "Answer", "Latency (ms)", "Tokens", "Sources"])
        for row, ans in enumerate(pr.get("answers", []), 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=ans.get("question", ""))
            ws.cell(row=row, column=3, value=ans.get("ground_truth", ""))
            ws.cell(row=row, column=4, value=ans.get("answer", ""))
            ws.cell(row=row, column=5, value=round(ans.get("latency_ms", 0), 1))
            ws.cell(row=row, column=6, value=ans.get("tokens_used", 0))
            ws.cell(row=row, column=7, value="\n---\n".join(ans.get("source_chunks", [])[:3]))
        for col, w in zip("ABCDEFG", [5, 40, 30, 40, 14, 10, 60]):
            ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Background executor
# ---------------------------------------------------------------------------

async def _execute_run(run_id: str, config: ExperimentConfig):
    from backend.services.experiment_executor import execute_experiment

    run = _active_runs[run_id]
    try:
        result = await execute_experiment(run_id, config, run)
        (RESULTS_DIR / f"{run_id}.json").write_text(result.model_dump_json(indent=2))
        cache = _load_cache()
        cache[_config_hash(config)] = run_id
        _save_cache(cache)
        logger.info("run=%s saved to disk", run_id)
    except Exception as exc:
        run.status = "error"
        run.message = str(exc)
        logger.error("run=%s failed: %s", run_id, exc, exc_info=True)
